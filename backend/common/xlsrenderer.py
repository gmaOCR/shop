from drf_excel.renderers import XLSXRenderer
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


class CustomXLSXRenderer(XLSXRenderer):
    def render(self, data, accepted_media_type=None, renderer_context=None):
        xlsx_data = super().render(data, accepted_media_type, renderer_context)
        workbook = load_workbook(BytesIO(xlsx_data))
        worksheet = workbook.active

        # Apply header styles
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()